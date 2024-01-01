#############################################################################################################
#Step 1: Load the xml file to extract the specific element and write the extracted data to excel file
#Step 2: Read the data from the ResultsTC file and produce an individual Excel file for each build
#Step 3: Compare build with the resultsByStatus file to retrieve the matching values and the test status
#Step 4: Plot1 display the total status count
#Step 5: Plot2 display the percentage of completion
# Developer: RHM
#############################################################################################################

#import modules
import subprocess
import os
import re
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.exceptions import InvalidFileException
import xlrd
from pathlib import Path
import itertools
from functools import reduce
import numpy as np
import matplotlib.pyplot as plt
import matplotlib as mpl
import glob
import xml.etree.ElementTree as ET
from lxml import etree
import plotly.express as px
import matplotlib.ticker as tkr
import logging
import math
import matplotlib.ticker as mtick
import string

#Read the attributes and extract the relevant element from the input xml by converting xml to excel file.
def xml_xl(input_folder,xml_folder):
    directory = input_folder
    logging.info('********** xml to excel File Conversion started ********** ')

    # Get the input XML file path
    xml_files = [f for f in os.listdir(directory) if f.endswith('.xml')]
    if not xml_files:
        logging.error('No XML file found in the input folder')
        return
    inp_name = os.path.join(directory, xml_files[0])
    logging.info('The Input File Locations is : ' + inp_name)
    parser = etree.XMLParser(remove_blank_text=True)
    root = ET.parse(inp_name,parser=parser).getroot()
    data=[]
    for child in root.findall('testsuite'):
        for each in child.findall('.//testcase'):
            version = each.find('.//version').text
            test_id = each.find('.//fullexternalid').text
            Test_Case = each.attrib['name']
            estimated_duration = each.find('.//estimated_exec_duration').text
            imp = each.find('.//importance').text
            for steps in each.findall('.//custom_fields'):
                for cs in steps.findall('.//custom_field'):
                    last = cs[-1].text
            data.append({'Test_case_ID': test_id,'Test_case_title': Test_Case,'Version':version,'Estimated_exec_[min]':estimated_duration,'Total_steps':last,'Priority':imp})
    df = pd.DataFrame(data)
    S = {'1':'Low','2':'Medium','3':'High'}
    df.columns = df.columns.str.strip()
    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    df['Priority'] = df['Priority'].map(S)
    df["Test_case_title"] = df["Test_case_ID"] + ':' + df["Test_case_title"]
    out_fname = xml_folder + 'TestSpecification_Testlink.xlsx'
    df.to_excel(out_fname)
    logging.info('The Excel Output file generated in : ' + out_fname)
    logging.info('********** xml to excel File Conversion Completed **********' + '\n')

#Create new files with the filtered columns from each build by reading the Result file
def build_calc(input_folder):
    directory = input_folder
    for filename in os.listdir(directory):
        if filename.startswith("resultsTC_") and filename.lower().endswith((".xlsx", ".xls",".XLS",".XLSX")):
            file_path = os.path.join(directory, filename)
            logging.info("The Input location to fetch resultsTC file is : " + directory)
            Exceldata = pd.read_excel(file_path, sheet_name='Worksheet', skiprows=4)
            df = pd.DataFrame(Exceldata)
            df = df.filter(regex=('Test Suite|Test Case|^Build|Date.*|Execution duration'))
            count = df.columns[df.columns.str.contains('Build')].value_counts().sum()
            dict_of_df = {}
            split_dfs = []
            for i in range(1, count + 1):
                dict_of_df["df_{}".format(i)] = df.iloc[:, -1 + (3 * i):(3 * i) + 2]
                dict_of_df["df_{}".format(i)]['Test Case'] = df['Test Case']
                dict_of_df["df_{}".format(i)]['Test Suite'] = df['Test Suite']
                f_n = dict_of_df["df_{}".format(i)].columns[0].replace(' ', '_')
                f_name = f_n.replace('+', '_').replace(':', '_').replace('__', '_')
                while '__' in f_name:
                    f_name = f_name.replace('__', '_')
                f_name += '.xlsx'
                logging.info("The Output file names are : " + f_name)
                dict_of_df["df_{}".format(i)].to_excel(os.path.join(directory, "Build_Input", f_name), index=False)
            logging.info("The filtered output files of each build is created in : " + directory+ '\n')

#Convert resultsByStatus xls file to xlsx file format
def xlstoxlsx(input_folder):
    folder_path = input_folder
    for xls_file in glob.glob(os.path.join(folder_path, "resultsByStatus_*.xls")):
        wb = xlrd.open_workbook(xls_file)
        sh = wb.sheet_by_index(0)
        xlsx_file = os.path.splitext(xls_file)[0] + '.xlsx'
        writer = openpyxl.Workbook()
        worksheet = writer.active
        worksheet.title = sh.name
        for row in range(1, sh.nrows + 1):
            for col in range(1, sh.ncols + 1):
                cell_value = sh.cell_value(row - 1, col - 1)
                worksheet.cell(row=row, column=col, value=cell_value)
        writer.save(xlsx_file)
        os.remove(xls_file)  # delete the original xls file

#Obtain the values of Passed, Partially passed, failed, blocked, clarification, not run, and total steps for each build by comparing the status file with each build file.
def test_report(input_folder,build_input,build_report,current_datetime):
    logging.info('Test report generation started')
    Status_file_path = Path(input_folder)
    Build_file_path = Path(build_input)
    Build_files = [x for x in Build_file_path.glob("**/*.xlsx") if x.name.__contains__("Build_")]
    Status_files = [x for x in Status_file_path.glob("**/*.xlsx") if x.name.__contains__("resultsByStatus_")]
    logging.info("Iterate over the status and build files to extract the status of each test case present in the build")

    for files in Build_files:
        for file in Status_files:
            df = pd.read_excel(files)
            df_builds = [col for col in df if col.startswith('Build')]
            df.loc[:, 'build'] = df_builds[0]
            df['Build'] = df['build'].str.split(' ', 1).str[1]
            df = df.drop('build', axis=1)
            data = load_workbook(file)
            ws = data.active
            val = data['Worksheet']['A1'].value
            status = val.rsplit(' ', 2)[0]
            df1 = pd.read_excel(file, skiprows=6)

            # List of required columns
            required_columns = ['Number of Clarification steps',
                                'Number of Partially Passed steps',
                                'Number of Passed steps',
                                'Number of Failed steps',
                                'Number of Blocked steps']

            for col in required_columns:
                if col not in df1.columns:
                    df1[col] = 0

            # Condition to fetch specific columns if string matches with status
            if 'Partially Passed' in status:
                df1 = df1[['Test Case', 'Build', 'Number of Partially Passed steps']]
                df_pp = df1.merge(df, on=['Test Case', 'Build'], how="right")
            elif 'Passed' in status:
                df1 = df1[['Test Case', 'Build', 'Number of Passed steps']]
                df_p = df1.merge(df, on=['Test Case', 'Build'], how="right")
            elif 'Failed' in status:
                df1 = df1[['Test Case', 'Build', 'Number of Failed steps']]
                df_f = df1.merge(df, on=['Test Case', 'Build'], how="right")
            elif 'Blocked' in status:
                df1 = df1[['Test Case', 'Build', 'Number of Blocked steps']]
                df_b = df1.merge(df, on=['Test Case', 'Build'], how="right")
            elif 'Clarification' in status:
                df1 = df1[['Test Case', 'Build', 'Number of Clarification steps']]
                df_c = df1.merge(df, on=['Test Case', 'Build'], how="right")
            elif 'Not Run' in status:
                df_read = pd.read_excel('../Reports/' + current_datetime + '/TestSpecificationReport/TestSpecification_Testlink.xlsx', index_col=0)
                df1 = df_read[['Test_case_title','Total_steps','Estimated_exec_[min]']]
                df_rn = df1.rename(columns={'Test_case_title':'Test Case'})
                df_nr = df_rn.merge(df, on=['Test Case'], how="right")
        data_frames = [df_pp,df_p,df_f,df_b,df_c,df_nr]

        # Remove the duplicated columns after merge
        df_merged = reduce(lambda left, right: pd.merge(left, right, on=['Test Case', 'Build'], how='outer',suffixes=('', '_x,_y,_z')), data_frames)
        df_merged.rename(columns={'Build': 'BD'}, inplace=True)
        df_merged.drop(df_merged.filter(regex='BD|_x,_y,_z$').columns, axis=1, inplace=True)
        df_merged['Number of Not Run Steps'] = df_merged['Total_steps'].where(df_merged.iloc[:, 2].str.contains('Not Run'))

        # Rearrange the column names
        df2 = df_merged[df_merged.columns[[5, 0, 2, 3, 4, 11, 6, 1, 7, 8, 9, 12, 10]]]
        df3 = df2.rename(columns={df2.columns[3]: 'Date', df2.columns[4]: 'Execution duration (min)'})

        # Convert date column to datetime object and format it as YYYY-mm-dd
        df3['Date'] = pd.to_datetime(df3['Date']).dt.strftime('%Y-%m-%d')

        valid_chars = "-_.()+ %s%s" % (string.ascii_letters, string.digits)
        Report_filename = ''.join(c if c in valid_chars else '_' for c in df3.columns[2]) + '.xlsx'
        Report_filename = Report_filename.replace(' ', '_')  # Replace spaces with underscores
        report_file_path = os.path.join(build_report, Report_filename)
        df3.to_excel(report_file_path, index=False)

def Supporting_func(input_folder,xml_folder,build_input,build_report,current_datetime):
    print(f"\nscript2:\nThe execution has begun to create a supporting files for the upcoming scripts. Please wait!..")
    xml_xl(input_folder,xml_folder)
    print(f"The xml to excel file conversion is completed. File generated in : {xml_folder}")
    build_calc(input_folder)
    print(f"Build Files have been created. File Generated in : {input_folder}"+'Build_Input/')
    xlstoxlsx(input_folder)
    print(f"The conversion of xls file type to .xlsx standard file format is completed. File generated in : {input_folder}")
    test_report(input_folder,build_input,build_report,current_datetime)
    print(f"The Report files have been created. File generated in : {build_report}")

# if __name__ == '__main__':
#     Supporting_func()